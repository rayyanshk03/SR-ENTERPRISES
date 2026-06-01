from fastapi import FastAPI, APIRouter
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List
import uuid
from datetime import datetime, timezone


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

EXCEL_PATH = "/Users/rayyanshaikh/SR-ENTERPRISES/Enquiries.xlsx"

def append_to_excel(lead_data: dict):
    # Ensure directory exists
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    
    # Check if file exists, if not create new with formatted headers
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enquiries"
        headers = [
            "Date & Time", "Source", "Name", "Phone Number", 
            "Company / School", "City", "Product Category", 
            "Quantity", "Message / Query", "Status", "Remarks"
        ]
        ws.append(headers)
        
        # Style headers (Dark Blue theme)
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        wb.save(EXCEL_PATH)
    
    # Load workbook and worksheet
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Enquiries"]
    
    # Get next row number
    next_row = ws.max_row + 1
    
    # Format current local time
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Prepare row data
    row_data = [
        now_str,
        lead_data.get("source", ""),
        lead_data.get("name", ""),
        lead_data.get("phone", ""),
        lead_data.get("company", ""),
        lead_data.get("city", ""),
        lead_data.get("category", ""),
        lead_data.get("quantity", ""),
        lead_data.get("message", ""),
        "Not Done",  # Default Status
        ""           # Empty Remarks by default
    ]
    
    ws.append(row_data)
    
    # Add validation dropdown for the 'Status' column (column 10)
    dv = DataValidation(type="list", formula1='"Not Done,In Progress,Done,Cancelled"', allow_blank=True)
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Select status from the list'
    dv.promptTitle = 'Status Options'
    
    ws.add_data_validation(dv)
    dv.add(ws.cell(row=next_row, column=10))
    
    # Dynamic styling for standard data row cells
    row_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=next_row, column=col_idx)
        cell.font = row_font
        # Center dates, sources, phones, quantities, statuses
        if col_idx in [1, 2, 4, 8, 10]:
            cell.alignment = center_align
        else:
            cell.alignment = left_align
            
    # Auto-adjust column widths for perfect readability
    for col in ws.columns:
        max_len = 0
        for cell in col:
            # Avoid counting formulas in data validation
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
        
    wb.save(EXCEL_PATH)

class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")  # Ignore MongoDB's _id field
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

class EnquiryCreate(BaseModel):
    source: str
    name: str
    phone: str
    company: str = ""
    city: str = ""
    category: str = ""
    quantity: str = ""
    message: str = ""

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/enquiry")
async def create_enquiry(input: EnquiryCreate):
    lead_dict = input.model_dump()
    try:
        append_to_excel(lead_dict)
    except Exception as e:
        logger.error(f"Failed to append to Excel sheet: {e}")
    return {"status": "success"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    # Exclude MongoDB's _id field from the query results
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()